from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
DATE = "10/08/2026"


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]
    row = 48

    values = {
        "B": DATE,
        "C": "Casa e limpeza",
        "D": "Refil de mop chenille branco com base plástica preta, pacote com cinco unidades para limpeza e troca.",
        "E": "Não identificada",
        "F": "Refil mop chenille pack 5",
        "G": "Novo",
        "H": "Não",
        "I": "Produto de limpeza; conferir encaixe, costura e integridade dos refis.",
        "J": "Disponível",
        "K": "A1",
        "N": 24.90,
        "O": 34.90,
        "P": 49.90,
        "Q": "Shopee nacional: Refil Mop Chenille Branco para limpeza R$24,90; Refil Mop Esfregão Chenille Premium R$34,90; Mop chenille kit com múltiplas unidades R$49,90.",
        "S": 35.00,
        "T": 35.00,
        "W": "Não",
        "X": "Não publicado",
        "Y": "Refil de mop chenille branco com base plástica preta, pacote com cinco unidades para limpeza e troca. Produto novo.",
        "AE": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    for col, value in values.items():
        ws[f"{col}{row}"] = value

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
