from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
DATE = "10/08/2026"


def fill_row(ws, row: int, data: dict[str, object]) -> None:
    for col, value in data.items():
        ws[f"{col}{row}"] = value
    ws[f"AE{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    fill_row(
        ws,
        35,
        {
            "B": DATE,
            "C": "Casa e limpeza",
            "D": "Calisul tira manchas com percarbonato de sódio, embalagem família, para uso em limpeza e remoção de manchas.",
            "E": "Calisul",
            "F": "Percarbonato de sódio tira manchas",
            "G": "Novo",
            "H": "Não",
            "I": "Produto em pó; conferir embalagem, lacre e peso antes de publicar.",
            "J": "Disponível",
            "K": "A1",
            "N": 22.98,
            "O": 30.69,
            "P": 49.50,
            "Q": "Shopee nacional: Percabonato De Sódio 100% Puro Calisul Tira Manchas Sem Cloro Roupas Brancas Ou Coloridas - ORIGINAL R$22,98-R$49,50; Percarbonato de Sódio Calisul - Limpeza Versátil e Eficaz R$30,69.",
            "S": 25.00,
            "T": 25.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Calisul tira manchas com percarbonato de sódio, indicado para limpeza doméstica e remoção de manchas. Produto novo.",
        },
    )

    fill_row(
        ws,
        36,
        {
            "B": DATE,
            "C": "Beleza e cuidados pessoais",
            "D": "Kit de bálsamo labial hidratante NORO com 4 unidades, em embalagem de conjunto para cuidados com os lábios.",
            "E": "NORO",
            "F": "Kit bálsamo labial 4 un",
            "G": "Novo",
            "H": "Não",
            "I": "Produto de higiene e beleza; conferir lacre e integridade das embalagens.",
            "J": "Disponível",
            "K": "A1",
            "N": 10.00,
            "O": 21.00,
            "P": 21.54,
            "Q": "Shopee nacional: Bálsamo labial hidratante Noro kit com 4 opções R$21,00; NORO Bálsamo Hidratante Labial R$10,00; Conjunto de Bálsamo Labial Hidratante NORO (8 peças) R$21,54.",
            "S": 20.00,
            "T": 20.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Kit de bálsamo labial hidratante NORO com 4 unidades, ideal para uso diário e cuidados com os lábios. Produto novo.",
        },
    )

    fill_row(
        ws,
        37,
        {
            "B": DATE,
            "C": "Beleza e cuidados pessoais",
            "D": "Base concreto endurecedora com formol para unhas, indicada para fortalecimento e acabamento de cuidado das unhas.",
            "E": "Top Beauty",
            "F": "Concreto endurecedor com formol",
            "G": "Novo",
            "H": "Não",
            "I": "Produto de uso cosmético; conferir validade, lacre e conservação da embalagem.",
            "J": "Disponível",
            "K": "A1",
            "N": 11.90,
            "O": 12.99,
            "P": 24.99,
            "Q": "Shopee nacional: Base Concreto Endurecedor Com Formol SOS unhas Top Beauty 7 ml R$12,99; Base Concreto Top Beauty Endurecedora de Unha com Formol R$11,90; KIT 03 UN Base Concreto Endurecedor Com Formol SOS unhas Top Beauty 7 ml R$24,99.",
            "S": 15.00,
            "T": 15.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Base concreto endurecedora com formol para unhas, produto novo para fortalecimento e cuidado estético.",
        },
    )

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
