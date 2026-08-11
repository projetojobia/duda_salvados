from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 31


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS030",
        "10/08/2026",
        "Casa e cozinha",
        "Garrafa de vidro para geladeira Mimo Style, capacidade 1L, com tampa inox e corpo transparente com textura em linhas. Ideal para agua, suco ou cha gelado.",
        "Mimo Style",
        "Garrafa Lines de vidro para geladeira 1L",
        "Novo",
        "Nao se aplica",
        "Produto sem mecanismo; conferir tampa, vidro e vedacao antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        19.24,
        27.90,
        59.50,
        "Base principal Shopee nacional: Kit 3 pecas Garrafa de Vidro 1 Litro com Alca Tampa Hermetica Geladeira Mimo Style R$59,50; Garrafa de Vidro 1 Litro Tampa Inox Agua Suco Geladeira R$39,90; Garrafa Lines de Vidro para Geladeira Mimostyle R$19,24 no pix em loja nacional. Produto equivalente por marca Mimo Style, capacidade 1L e tampa inox.",
        '=IF(OR(R31="";T31="");"";R31-T31)',
        20.00,
        20.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Garrafa de vidro para geladeira Mimo Style, 1L, com tampa inox e corpo transparente com textura em linhas. Indicada para agua, suco e cha gelado. Produto novo.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe alta saida por item domestico de baixo ticket: S=N*0,90, arredondado comercialmente para R$20,00. Preco final definido automaticamente conforme nova regra. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS030 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
