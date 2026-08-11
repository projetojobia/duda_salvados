from pathlib import Path

from openpyxl import load_workbook


PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")


def main() -> None:
    wb = load_workbook(PATH)
    ws = wb["Produtos"]
    for r in range(26, 33):
        ws[f"R{r}"] = f'=IF(N{r}="";"";N{r}*50%)'
        ws[f"U{r}"] = f'=IF(OR(R{r}="";T{r}="");"";T{r}-R{r})'
        ws[f"V{r}"] = f'=IF(OR(T{r}="";T{r}=0;U{r}="");"";U{r}/T{r})'
    wb.save(PATH)


if __name__ == "__main__":
    main()
