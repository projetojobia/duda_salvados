from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 28


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS027",
        "10/08/2026",
        "Casa e cozinha",
        "Tábua de corte e servir em madeira, formato retangular, com alça/pegador vazado, acabamento marrom natural e bordas trabalhadas. Ideal para churrasco, frios, petiscos e cozinha.",
        "Nao identificada",
        "Tabua de corte e servir em madeira com alca",
        "Novo",
        "Nao se aplica",
        "Produto sem mecanismo; conferir medidas, acabamento e estado da madeira antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        35.96,
        41.19,
        59.00,
        "Base principal Shopee nacional: Tábua de Corte com Alça em Madeira Natural 49cm para Cozinha e Servir R$41,19; Tábua de Corte Madeira Teca Maciça 30x20cm Com Canaleta e Alça R$35,96; Tábua de Corte e Servir em Madeira Maciça R$59,00. Produto equivalente por formato, uso na cozinha/churrasco e presença de alça.",
        '=IF(OR(R28="";T28="");"";R28-T28)',
        35.00,
        35.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Tábua de corte e servir em madeira com alça, acabamento marrom natural e formato retangular. Indicada para churrasco, frios, petiscos e preparo na cozinha.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe media saida por item de cozinha em madeira: S=N*0,75, arredondado comercialmente para R$35,00. Preco final definido automaticamente conforme nova regra. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS027 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
