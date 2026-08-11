from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
ROW = 33


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    values = {
        "B": "10/08/2026",
        "C": "Ferramentas e melhorias para casa",
        "D": "Extensão elétrica preta de 15 metros, uso geral, com plugue e tomadas múltiplas, cabo reforçado.",
        "E": "Não identificada",
        "F": "Extensão elétrica 15 m",
        "G": "Novo",
        "H": "Não",
        "I": "Produto sem mecanismo; conferir cabo, plugue, tomadas e continuidade antes de publicar.",
        "J": "Disponível",
        "K": "A1",
        "N": 39.90,
        "O": 56.99,
        "P": 66.04,
        "Q": (
            "Shopee nacional: Extensão Elétrica 15 Metros 10A Cabo PP 2x1,0 Reforçada Certificada Anti-chama "
            "R$39,90; Prolongador Extensão Elétrica 15 Metros 10a Cabo Pp2x2,5mm Plug Bipolar Injetado "
            "R$56,99; Extensão Elétrica 15 Metros 10a Cabo Pp Reforçada Cor Preto 127V/220V R$61,65."
        ),
        "S": 55.00,
        "T": 55.00,
        "W": "Não",
        "X": "Não publicado",
        "Y": (
            "Extensão elétrica preta de 15 metros, prática para uso doméstico e geral. Produto novo, com cabo "
            "reforçado e conjunto de tomadas múltiplas."
        ),
        "AE": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    for col, value in values.items():
        ws[f"{col}{ROW}"] = value

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
