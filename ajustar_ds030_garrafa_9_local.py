from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 31


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]
    ws[f"T{ROW}"] = 9
    ws[f"AE{ROW}"] = "10/08/2026 14:35:00"
    wb.save(WORKBOOK_PATH)
    print("DS030 atualizado para R$9,00 na planilha local.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
