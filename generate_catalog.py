from __future__ import annotations

import csv
import json
import re
import shutil
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, ImageOps


WORKBOOK = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2_dashboard_executivo.xlsx")
PHOTO_REPORT = Path(r"C:\Users\User\duda\photo_rename_report.csv")
PHOTO_OVERRIDES = Path(r"C:\Users\User\duda\catalog_photo_overrides.json")
PRODUCT_OVERRIDES = Path(r"C:\Users\User\duda\catalog_product_overrides.json")
MANUAL_MEDIA_DIR = Path(r"C:\Users\User\duda\catalog_manual_media")
PUBLIC_DIR = Path(r"C:\Users\User\duda\public")
ASSETS_DIR = PUBLIC_DIR / "assets" / "products"
DATA_PATH = PUBLIC_DIR / "catalog.json"
IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp"}
VIDEO_EXTENSIONS = {".mp4", ".mov", ".webm"}


COLS = {
    "code": 1,
    "category": 3,
    "title": 4,
    "brand": 5,
    "model": 6,
    "condition": 7,
    "tested": 8,
    "working": 9,
    "status": 10,
    "location": 11,
    "price": 20,
    "catalog_status": 24,
    "ad_description": 25,
    "quantity": 32,
}

PUBLIC_STATUSES = {"Disponível", "Pronto para publicar", "Publicado"}


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text[:90] or "produto"


def cell(row: tuple[Any, ...], key: str) -> Any:
    index = COLS[key] - 1
    if index >= len(row):
        return ""
    return row[index] or ""


def load_photo_map() -> dict[str, list[Path]]:
    photos: dict[str, list[Path]] = defaultdict(list)
    if not PHOTO_REPORT.exists():
        return photos
    with PHOTO_REPORT.open("r", encoding="utf-8-sig", newline="") as f:
        for item in csv.DictReader(f):
            if item.get("Tipo") != "PRODUTO":
                continue
            code = (item.get("Codigo") or "").strip()
            path = Path(item.get("DestinoArquivo") or "")
            if code and path.exists():
                photos[code].append(path)
    if MANUAL_MEDIA_DIR.exists():
        for code_dir in MANUAL_MEDIA_DIR.iterdir():
            if not code_dir.is_dir():
                continue
            for path in sorted(code_dir.iterdir()):
                if path.suffix.lower() in IMAGE_EXTENSIONS | VIDEO_EXTENSIONS:
                    photos[code_dir.name].append(path)
    return {code: sorted(paths) for code, paths in photos.items()}


def load_photo_overrides() -> dict[str, dict[str, Any]]:
    if not PHOTO_OVERRIDES.exists():
        return {}
    return json.loads(PHOTO_OVERRIDES.read_text(encoding="utf-8"))


def load_product_overrides() -> dict[str, dict[str, Any]]:
    if not PRODUCT_OVERRIDES.exists():
        return {}
    return json.loads(PRODUCT_OVERRIDES.read_text(encoding="utf-8"))


def apply_photo_overrides(code: str, photos: list[Path], overrides: dict[str, dict[str, Any]]) -> list[Path]:
    override = overrides.get(code, {})
    hidden = set(override.get("hidden") or [])
    preferred_order = list(override.get("order") or [])
    primary = override.get("primary") or ""
    by_name = {photo.name: photo for photo in photos if photo.name not in hidden}
    ordered: list[Path] = []

    if primary and primary in by_name:
        ordered.append(by_name.pop(primary))

    for name in preferred_order:
        if name in by_name:
            ordered.append(by_name.pop(name))

    ordered.extend(by_name[name] for name in sorted(by_name))
    return ordered


def optimize_image(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with Image.open(source) as img:
        img = ImageOps.exif_transpose(img)
        img.thumbnail((1200, 1200), Image.Resampling.LANCZOS)
        if img.mode not in {"RGB", "L"}:
            img = img.convert("RGB")
        img.save(target, "WEBP", quality=78, method=6)


def publish_video(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)


def build_catalog() -> dict[str, Any]:
    photos_by_code = load_photo_map()
    photo_overrides = load_photo_overrides()
    product_overrides = load_product_overrides()
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Produtos"]

    products: list[dict[str, Any]] = []
    missing_photos: list[str] = []
    skipped_sold = 0

    for row in ws.iter_rows(min_row=2, max_row=301, min_col=1, max_col=32, values_only=True):
        code = str(cell(row, "code")).strip()
        if not code or not code.startswith("DS"):
            continue
        has_data = any(value not in (None, "") for value in row[1:])
        if not has_data:
            continue

        status = str(cell(row, "status")).strip()
        if status not in PUBLIC_STATUSES:
            skipped_sold += 1
            continue

        default_title = str(cell(row, "model") or cell(row, "title")).strip()
        title = str(product_overrides.get(code, {}).get("title") or default_title).strip()
        slug = slugify(f"{code}-{title}")
        image_urls: list[str] = []
        media_items: list[dict[str, str]] = []
        curated_photos = apply_photo_overrides(code, photos_by_code.get(code, []), photo_overrides)
        for index, source in enumerate(curated_photos, start=1):
            suffix = source.suffix.lower()
            if suffix in IMAGE_EXTENSIONS:
                image_name = f"{slug}-{index:02d}.webp"
                target = ASSETS_DIR / image_name
                optimize_image(source, target)
                url = f"/assets/products/{image_name}"
                image_urls.append(url)
                media_items.append({"type": "image", "url": url})
            elif suffix in VIDEO_EXTENSIONS:
                video_name = f"{slug}-{index:02d}{suffix}"
                target = ASSETS_DIR / video_name
                publish_video(source, target)
                media_items.append({"type": "video", "url": f"/assets/products/{video_name}"})

        if not image_urls:
            missing_photos.append(code)

        products.append(
            {
                "code": code,
                "slug": slug,
                "title": title,
                "description": str(cell(row, "ad_description") or cell(row, "title")).strip(),
                "category": str(cell(row, "category")).strip() or "Sem categoria",
                "brand": str(cell(row, "brand")).strip(),
                "model": str(cell(row, "model")).strip(),
                "condition": str(cell(row, "condition")).strip(),
                "tested": str(cell(row, "tested")).strip(),
                "working": str(cell(row, "working")).strip(),
                "status": status,
                "catalogStatus": str(cell(row, "catalog_status")).strip(),
                "location": str(cell(row, "location")).strip(),
                "price": str(cell(row, "price")).strip(),
                "quantity": str(cell(row, "quantity")).strip() or "1",
                "images": image_urls,
                "media": media_items,
                "whatsAppText": f"Ola, tenho interesse no {code} - {title}",
            }
        )

    categories = sorted({product["category"] for product in products})
    total_value = sum(parse_brl(product["price"]) * parse_int(product["quantity"]) for product in products)
    return {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "source": {
            "spreadsheetId": "1kqLutUpgQwwnJgHmR7wrJ97zvQR8QZGRB6QyymM4GPU",
            "workbook": WORKBOOK.name,
        },
        "summary": {
            "products": len(products),
            "categories": len(categories),
            "totalPotential": total_value,
            "missingPhotos": missing_photos,
            "skippedNonPublic": skipped_sold,
        },
        "categories": categories,
        "products": products,
    }


def parse_brl(value: str) -> float:
    cleaned = value.replace("R$", "").replace(".", "").replace(",", ".")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def parse_int(value: str) -> int:
    try:
        return max(1, int(float(value.replace(",", "."))))
    except ValueError:
        return 1


def main() -> int:
    if ASSETS_DIR.exists():
        shutil.rmtree(ASSETS_DIR)
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    catalog = build_catalog()
    DATA_PATH.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(catalog["summary"], ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
