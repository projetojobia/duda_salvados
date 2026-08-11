from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 32


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS031",
        "10/08/2026",
        "Casa e organizacao",
        "Kit com 12 caixas organizadoras transparentes pequenas, empilhaveis, estilo multiuso para armario, gaveta, cozinha ou itens diversos.",
        "Nao identificada",
        "Caixas organizadoras transparentes kit 12 unidades",
        "Novo",
        "Nao se aplica",
        "Produto sem mecanismo; conferir quantidade, acabamento e estado geral antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        19.80,
        25.00,
        49.90,
        "Base principal Shopee nacional: Caixa Organizadora Plastica 30 Litros Transparente Com Tampa Trava Empilhavel Multiuso Casa Quarto R$37,99; Caixa Organizadora Multiuso Transparente Empilhavel C/ Tampa e Trava R$17,99-R$69,99; kit 12 caixas organizadoras transparentes/acrilicas semelhantes em torno de R$155,99, que nao e base direta por ser kit de sapato. Produto equivalente forte por formato transparente, uso organizador e empilhavel.",
        '=IF(OR(R32="";T32="");"";R32-T32)',
        60.00,
        60.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Kit com 12 caixas organizadoras transparentes pequenas, empilhaveis e multiuso. Ideal para armario, gaveta, cozinha ou organizacao de itens diversos. Produto novo.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe alta saida por produto utilitario organizador: S=N*0,90, mas o preco final foi solicitado pelo usuario em R$60,00. Preco final definido pelo usuario. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS031 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
