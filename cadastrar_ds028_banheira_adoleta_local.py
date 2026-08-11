from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 29


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS028",
        "10/08/2026",
        "Mãe e bebê",
        "Banheira de bebê Adoleta, cor rosa, capacidade 15L, produto plástico para banho infantil. Etiqueta indica modelo RSBB e cor rosa bebe.",
        "Adoleta",
        "Banheira de bebê 15L rosa RSBB",
        "Novo",
        "Nao se aplica",
        "Produto sem mecanismo; conferir integridade, tampa/acabamento e capacidade antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        33.91,
        35.69,
        62.99,
        "Base principal Shopee nacional: BANHEIRA DE BEBÊ ADOLETA 20 LITROS AZUL ROSA R$33,91; Banheira Infantil Adoleta Aconchego 22 Litros Azul ou Rosa R$62,99; Banheira Bebê Infantil rígida/compacta similares entre R$33,91 e R$62,99. Produto equivalente por marca Adoleta, cor rosa e uso infantil de banho.",
        '=IF(OR(R29="";T29="");"";R29-T29)',
        35.00,
        35.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Banheira de bebê Adoleta, cor rosa, 15L, modelo plástico para banho infantil. Produto novo; ideal para uso doméstico no banho do bebê.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe media saida por item infantil de banho: S=N*0,75, arredondado comercialmente para R$35,00. Preco final definido automaticamente conforme nova regra. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS028 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
