from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
DATE = "10/08/2026"


def fill_row(ws, row: int, data: dict[str, object]) -> None:
    for col, value in data.items():
        ws[f"{col}{row}"] = value
    ws[f"AE{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    fill_row(
        ws,
        38,
        {
            "B": DATE,
            "C": "Beleza e cuidados com a pele",
            "D": "Creme facial com retinol, colágeno e ácido hialurônico, indicado para uso diário e anti-idade.",
            "E": "OUHOE",
            "F": "Retinol Colágeno Complex + HA",
            "G": "Novo",
            "H": "Não",
            "I": "Produto cosmético; conferir lacre, validade e integridade da embalagem.",
            "J": "Disponível",
            "K": "A1",
            "N": 26.98,
            "O": 32.90,
            "P": 39.99,
            "Q": "Shopee nacional: Retinol Colágeno Complex + Ácido Hialurônico Anti Rugas Anti Envelhecimento R$32,90; Colágeno Complex Ácido Hialurônico Creme Anti Rugas R$39,99; Creme Hidratante e Rejuvenescedor com Ácido Hialurônico, Vitamina A (Retinol) e Óleo de Rosa Mosqueta 30g R$54,00.",
            "S": 35.00,
            "T": 35.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Creme facial com retinol, colágeno e ácido hialurônico para hidratação e cuidado anti-idade. Produto novo.",
        },
    )

    fill_row(
        ws,
        39,
        {
            "B": DATE,
            "C": "Automotivo e limpeza",
            "D": "Removedor de ferrugem Wurth 250 ml, indicado para limpeza, descontaminação e remoção de ferrugem em superfícies metálicas.",
            "E": "Wurth",
            "F": "Removedor de ferrugem 250 ml",
            "G": "Novo",
            "H": "Não",
            "I": "Produto químico; conferir lacre, volume e conservação antes da publicação.",
            "J": "Disponível",
            "K": "A1",
            "N": 27.03,
            "O": 28.90,
            "P": 32.00,
            "Q": "Shopee nacional: Removedor De Ferrugem Limpa Chassi Max Wurth 250ml R$29,38; Removedor De Ferrugem Limpa Oxidação Fosfatiza Wurth 250ml R$28,90; Removedor De Ferrugem Limpa Oxidação Fosfatiza Wurth 250ml V Verde R$32,00.",
            "S": 30.00,
            "T": 30.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Removedor de ferrugem Wurth 250 ml para limpeza e descontaminação de superfícies metálicas. Produto novo.",
        },
    )

    fill_row(
        ws,
        40,
        {
            "B": DATE,
            "C": "Saúde e suplementos",
            "D": "Suplemento termogênico Dimethylx Under Labz, frasco com 60 cápsulas, indicado para apoio à performance e energia.",
            "E": "Under Labz",
            "F": "Dimethylx 60 cápsulas",
            "G": "Novo",
            "H": "Não",
            "I": "Produto suplementar; conferir lacre, validade e rótulo antes de publicar.",
            "J": "Disponível",
            "K": "A1",
            "N": 66.43,
            "O": 69.93,
            "P": 99.90,
            "Q": "Shopee nacional: Termogenico Dimethylex Fat Burner 60 Cápsulas - Under Labz R$66,43 / R$69,93 com cupom; variações nacionais semelhantes em suplementos de performance.",
            "S": 70.00,
            "T": 70.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Suplemento termogênico Dimethylx Under Labz com 60 cápsulas, indicado para apoio à performance e energia. Produto novo.",
        },
    )

    fill_row(
        ws,
        41,
        {
            "B": DATE,
            "C": "Beleza e cuidados com o cabelo",
            "D": "Máscara capilar Arvensis Cachos Naturais 2x1, frasco de 500 g, para hidratação e nutrição de cabelos crespos e crespíssimos.",
            "E": "Arvensis",
            "F": "Cachos Naturais 500 g",
            "G": "Novo",
            "H": "Não",
            "I": "Produto cosmético; conferir lacre, validade e integridade da embalagem.",
            "J": "Disponível",
            "K": "A1",
            "N": 64.50,
            "O": 94.18,
            "P": 110.80,
            "Q": "Shopee nacional: Máscara Capilar Arvensis Cachos Naturais 2x1 500g R$94,18 / R$110,80; referência nacional forte para tratamento capilar de cachos.",
            "S": 90.00,
            "T": 90.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Máscara capilar Arvensis Cachos Naturais 2x1, ideal para hidratação e nutrição de cabelos crespos e crespíssimos. Produto novo.",
        },
    )

    fill_row(
        ws,
        42,
        {
            "B": DATE,
            "C": "Beleza e cuidados com a pele",
            "D": "Colágeno hidrolisado Covitta Beauty com ácido hialurônico e vitamina C, suplemento alimentar em pó de 250 g.",
            "E": "Covitta Beauty",
            "F": "Colágeno hidrolisado 250 g",
            "G": "Novo",
            "H": "Não",
            "I": "Produto suplementar; conferir lacre, validade e conservação da embalagem.",
            "J": "Disponível",
            "K": "A1",
            "N": 29.99,
            "O": 29.99,
            "P": 29.99,
            "Q": "Shopee nacional: Colágeno 250g Hidrolisado com Ácido Hialurônico, Biotina, Retinol e Peptídeos Bioativos tipo Verisol R$29,99.",
            "S": 30.00,
            "T": 30.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Colágeno hidrolisado Covitta Beauty com ácido hialurônico e vitamina C, suplemento alimentar em pó de 250 g. Produto novo.",
        },
    )

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
