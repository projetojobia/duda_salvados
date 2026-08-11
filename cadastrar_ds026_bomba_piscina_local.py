from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 27


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS026",
        "10/08/2026",
        "Casa e jardim",
        "Motobomba de piscina Sibrape/WEG BPF-050, 1/2 CV, motor WEG, eixo inox, bivolt 110-127/220-254V, modelo BPF 050 W01 CB. Produto fotografado sem caixa.",
        "Sibrape",
        "BPF-050 W01 CB",
        "Usado",
        "Nao se aplica",
        "Conferir motor, conexoes, rotor e funcionamento eletrico antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        523.17,
        680.29,
        1129.44,
        "Base principal Shopee nacional: Motobomba Piscina 1/2cv Pré Filtro Motor Silencioso Cascata 220v/127v Sistema De Filtragem R$523,17; Motobomba Piscina 1/2CV Motor WEG Bivolt 127/220V Alta Vazão R$680,29; Motobomba Sibrape 1/2CV BPF-050 - Eixo Inox 50/60Hz 110-127/220-254V R$1.129,44. Produto equivalente por potência 1/2 CV, motor WEG/Sibrape e uso para piscina.",
        '=IF(OR(R27="";T27="");"";R27-T27)',
        500.00,
        500.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Motobomba de piscina Sibrape/WEG BPF-050, 1/2 CV, bivolt, eixo inox, indicada para filtragem e recirculacao de agua. Produto usado, sem caixa. Conferir funcionamento eletrico e conexoes antes da venda.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe media saida por equipamento tecnico usado e sem teste: S=N*0,75, arredondado comercialmente para R$500,00. Preco final definido automaticamente conforme nova regra. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS026 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
