from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2_dashboard_executivo.xlsx")
PRODUCT_OVERRIDES_PATH = Path(r"C:\Users\User\duda\catalog_product_overrides.json")
SPREADSHEET_ID = "1kqLutUpgQwwnJgHmR7wrJ97zvQR8QZGRB6QyymM4GPU"
EXPECTED_TITLE = "Duda Salvados - Base Oficial Restaurada"
GOOGLE_SCRIPTS = Path(
    r"C:\Users\User\AppData\Local\hermes\skills\productivity\google-workspace\scripts"
)

sys.path.insert(0, str(GOOGLE_SCRIPTS))

from google_api import build_service  # noqa: E402


def convert_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    return value


def convert_formula_for_google(value: str) -> str:
    replacements = {
        "IFERROR": "SEERRO",
        "COUNTIF": "CONT.SE",
        "SUMPRODUCT": "SOMARPRODUTO",
        "SUMIFS": "SOMASES",
        "SUM": "SOMA",
        "MAX": "MÁXIMO",
        "OR": "OU",
        "IF": "SE",
    }
    formula = value
    for english, portuguese in sorted(replacements.items(), key=lambda item: len(item[0]), reverse=True):
        formula = formula.replace(f"{english}(", f"{portuguese}(")
    return formula.replace(",", ";")


def sheet_values(ws) -> list[list[Any]]:
    data: list[list[Any]] = []
    for row in ws.iter_rows():
        data.append([convert_value(cell.value) for cell in row])
    while data and all(value == "" for value in data[-1]):
        data.pop()
    if not data:
        return [[""]]
    max_cols = max(len(row) for row in data)
    return [row + [""] * (max_cols - len(row)) for row in data]


def sheet_bounds(ws) -> tuple[int, int]:
    last_row = 1
    last_col = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                last_row = max(last_row, cell.row)
                last_col = max(last_col, cell.column)
    return last_row, last_col


def load_product_overrides() -> dict[str, dict[str, Any]]:
    if not PRODUCT_OVERRIDES_PATH.exists():
        return {}
    return json.loads(PRODUCT_OVERRIDES_PATH.read_text(encoding="utf-8-sig"))


def apply_product_overrides_to_raw_values(sheet_name: str, raw_values: list[list[Any]]) -> None:
    if sheet_name != "Produtos" or len(raw_values) < 2:
        return
    overrides = load_product_overrides()
    if not overrides:
        return

    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    for row in raw_values[1:]:
        if not row:
            continue
        code = str(row[0] or "").strip()
        override = overrides.get(code)
        if not override:
            continue

        while len(row) < 32:
            row.append("")

        if override.get("title"):
            row[5] = override["title"]  # Modelo
        if override.get("price") not in (None, ""):
            row[19] = override["price"]  # Preco definido
        if override.get("referencePrice") not in (None, ""):
            row[14] = override["referencePrice"]  # Preco medio IA
            row[15] = override["referencePrice"]  # Preco alto IA
        if override.get("sold"):
            row[9] = "Vendido"  # Status interno
            row[23] = "Vendido"  # Status catalogo
            if not row[25]:
                row[25] = now  # Data venda
        elif override.get("reserved"):
            row[9] = "Reservado"  # Status interno
            row[23] = "Reservado"  # Status catalogo
        elif override.get("hidden"):
            row[23] = "Oculto"  # Status catalogo
        row[30] = now  # Ultima atualizacao


def build_payload(workbook_path: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    wb = load_workbook(workbook_path, data_only=False)
    raw_data = []
    formula_data = []
    for ws in wb.worksheets:
        values = sheet_values(ws)
        raw_values = [
            ["" if isinstance(value, str) and value.startswith("=") else value for value in row]
            for row in values
        ]
        apply_product_overrides_to_raw_values(ws.title, raw_values)
        last_row, last_col = sheet_bounds(ws)
        width = max(last_col, max((len(row) for row in values), default=1))
        height = max(last_row, len(values))
        raw_data.append(
            {
                "range": f"{ws.title}!A1:{get_column_letter(width)}{height}",
                "values": raw_values,
            }
        )
        for row_index, row in enumerate(values, start=1):
            for col_index, value in enumerate(row, start=1):
                if isinstance(value, str) and value.startswith("="):
                    formula_data.append(
                        {
                            "range": f"{ws.title}!{get_column_letter(col_index)}{row_index}",
                            "values": [[convert_formula_for_google(value)]],
                        }
                    )
    wb.close()
    return raw_data, formula_data


def verify_destination(service) -> str:
    metadata = service.spreadsheets().get(spreadsheetId=SPREADSHEET_ID, fields="properties/title").execute()
    title = metadata.get("properties", {}).get("title", "")
    if title != EXPECTED_TITLE:
        raise RuntimeError(f"Destino inesperado: {title!r}. Esperado: {EXPECTED_TITLE!r}")
    return title


def main() -> int:
    parser = argparse.ArgumentParser(description="Sincroniza a planilha local oficial com o Google Sheets restaurado.")
    parser.add_argument("--dry-run", action="store_true", help="Mostra o que seria enviado sem alterar o Google Sheets.")
    args = parser.parse_args()

    if not WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Planilha local nao encontrada: {WORKBOOK_PATH}")

    service = build_service("sheets", "v4")
    title = verify_destination(service)
    raw_data, formula_data = build_payload(WORKBOOK_PATH)

    total_cells = sum(len(item["values"]) * max((len(row) for row in item["values"]), default=0) for item in raw_data)
    print(f"Destino: {title}")
    print(f"Arquivo local: {WORKBOOK_PATH.name}")
    print(f"Abas: {len(raw_data)}")
    print(f"Celulas estimadas: {total_cells}")
    print(f"Formulas: {len(formula_data)}")

    if args.dry_run:
        print("Dry-run: nenhuma alteracao enviada.")
        return 0

    raw_result = service.spreadsheets().values().batchUpdate(
        spreadsheetId=SPREADSHEET_ID,
        body={"valueInputOption": "RAW", "data": raw_data},
    ).execute()
    formula_result = {}
    if formula_data:
        formula_result = service.spreadsheets().values().batchUpdate(
            spreadsheetId=SPREADSHEET_ID,
            body={"valueInputOption": "USER_ENTERED", "data": formula_data},
        ).execute()
    print({"raw": raw_result, "formulas": formula_result})
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
