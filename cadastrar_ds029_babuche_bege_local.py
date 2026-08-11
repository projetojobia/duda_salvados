from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 30


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS029",
        "10/08/2026",
        "Calcados",
        "Par de babuches/crocs bege claro, estilo confortavel com furos frontais e tiras traseiras, indicado para uso casual, praia ou dia a dia.",
        "Nao identificada",
        "Babuche/crocs bege adulto",
        "Novo",
        "Nao se aplica",
        "Produto sem mecanismo; conferir par, tamanho e estado geral antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        24.70,
        34.90,
        45.00,
        "Base principal Shopee nacional: Sandalia Babuche unissex Leve e Confortavel Adulto e Infantil R$24,70; Babuche Estiloso Unissex – Sandalia Anatomica Leve R$34,90; Babuche Masculino e Feminino Conforto De Calcar Leve R$39,90. Produto equivalente por estilo babuche/crocs bege e uso casual.",
        '=IF(OR(R30="";T30="");"";R30-T30)',
        25.00,
        25.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Par de babuches/crocs bege claro com furos frontais e tira traseira. Modelo confortavel para uso diario, praia ou casa. Produto novo; conferir par e tamanho antes da entrega.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe alta saida por item barato e casual: S=N*0,90, arredondado comercialmente para R$25,00. Preco final definido automaticamente conforme nova regra. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS029 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
