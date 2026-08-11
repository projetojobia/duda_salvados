from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
DATE = "10/08/2026"


def fill_row(ws, row: int, data: dict[str, object]) -> None:
    for col, value in data.items():
        ws[f"{col}{row}"] = value
    ws[f"AE{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    fill_row(
        ws,
        46,
        {
            "B": DATE,
            "C": "Cama, mesa e banho",
            "D": "Jogo de cama solteiro com 4 peças em microfibra 100% poliéster, contendo edredom dupla face, lençol de baixo, fronha e porta-travesseiro.",
            "E": "Não identificada",
            "F": "Jogo de cama solteiro 4 peças",
            "G": "Novo",
            "H": "Não",
            "I": "Produto têxtil; conferir costura, medidas, zíper e acabamento do kit.",
            "J": "Disponível",
            "K": "A1",
            "N": 39.90,
            "O": 49.90,
            "P": 79.90,
            "Q": "Shopee nacional: Jogo de Cama Solteiro 4 Peças Microfibra 100% Poliéster R$39,90; Kit Cama Solteiro 4 Peças Microfibra Estampado R$49,90; Jogo de Cama Solteiro com Edredom Dupla Face R$79,90.",
            "S": 40.00,
            "T": 40.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Jogo de cama solteiro com 4 peças em microfibra 100% poliéster, com edredom dupla face, lençol, fronha e porta-travesseiro. Produto novo.",
        },
    )

    fill_row(
        ws,
        47,
        {
            "B": DATE,
            "C": "Automotivo",
            "D": "Par de palhetas flexíveis para para-brisa, tamanhos 22 e 26 polegadas, modelos SW1908 e SW1917.",
            "E": "SW",
            "F": "Palhetas 22/26 polegadas",
            "G": "Novo",
            "H": "Não",
            "I": "Produto automotivo; conferir encaixe, borracha e integridade das hastes.",
            "J": "Disponível",
            "K": "A1",
            "N": 18.90,
            "O": 24.90,
            "P": 35.00,
            "Q": "Shopee nacional: Palheta de Silicone Flexível Frontal 26'' SW1917 R$18,90; Palheta de Silicone Flexível Frontal 22'' SW1908 R$24,90; Par de palhetas universais 22/26'' R$35,00.",
            "S": 25.00,
            "T": 25.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Par de palhetas flexíveis para para-brisa, tamanhos 22 e 26 polegadas, modelos SW1908 e SW1917. Produto novo.",
        },
    )

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
