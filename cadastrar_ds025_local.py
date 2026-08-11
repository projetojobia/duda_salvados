from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET = "Produtos"
ROW = 26


def main() -> int:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET]

    values = [
        "DS025",
        "10/08/2026",
        "Beleza e cuidados pessoais",
        "Esmalteira/expositor de esmaltes de parede, cor branca com estrutura marrom, modelo com cinco prateleiras para organizacao e exposicao de esmaltes.",
        "Nao identificada",
        "Esmalteiro expositor de esmaltes 5 prateleiras branco/marrom",
        "Novo",
        "Nao se aplica",
        "Produto sem mecanismo; conferir medidas, prateleiras e acabamento antes de publicar/entregar.",
        "Disponível",
        "A1",
        "",
        "",
        28.90,
        39.90,
        49.90,
        "Base principal Shopee nacional: Expositor de esmalte p/ manicure 200 esmaltes nicho de parede MDF R$39,88; Esmalteiro Expositor 5 Prateleiras Parede Branco Esmalte Gel Unhas Cílios Nails Organizador MDF R$49,98; Esmalteira/organizadores de parede similares entre R$28,90 e R$49,90. Produto equivalente por uso, formato de parede, cor branca e foco em esmaltes.",
        '=IF(OR(R26="";T26="");"";R26-T26)',
        35.00,
        35.00,
        "",
        "",
        "Nao",
        "Nao publicado",
        "Esmalteira/expositor de esmaltes de parede branco com estrutura marrom e cinco prateleiras. Ideal para organizar e expor esmaltes, tintas e itens pequenos de manicure. Produto novo.",
        "",
        "",
        "",
        "",
        "Fotos recebidas no Codex; ainda nao enviadas ao Drive. Base Shopee nacional por equivalente forte. Classe alta saida por item de organizacao para manicure: S=N*0,90, arredondado comercialmente para R$35,00. Preco final definido automaticamente conforme nova regra. Status inicial Disponivel.",
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    ]

    for idx, value in enumerate(values, start=1):
        ws.cell(row=ROW, column=idx).value = value

    wb.save(WORKBOOK_PATH)
    print(f"Gravado DS025 na linha {ROW}.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
