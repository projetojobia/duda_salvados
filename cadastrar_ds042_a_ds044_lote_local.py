from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
DATE = "10/08/2026"


def fill_row(ws, row: int, data: dict[str, object]) -> None:
    for col, value in data.items():
        ws[f"{col}{row}"] = value
    ws[f"AE{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    fill_row(
        ws,
        43,
        {
            "B": DATE,
            "C": "Alimentos para pets",
            "D": "Pacote com 15 sachês de alimento úmido Purina Friskies para gatos filhotes, sabor carne ao molho, 85 g cada sachê.",
            "E": "Purina Friskies",
            "F": "Sachês Friskies filhotes carne 15 un",
            "G": "Novo",
            "H": "Não",
            "I": "Produto alimentício para pets; conferir integridade da caixa, validade e conservação.",
            "J": "Disponível",
            "K": "A1",
            "N": 38.90,
            "O": 45.89,
            "P": 62.47,
            "Q": "Shopee nacional: Pack Purina Friskies Carne ao Molho C/15 Sachês de 85g R$51,91; Kit com 15 Sachês Friskies Filhotes 85g Sabor Carne ao molho R$54,00; Ração Friskies para Gatos Filhotes Sabor Carne ao Molho Sachê 85g - Embalagem com 15 Unidades R$62,47.",
            "S": 45.00,
            "T": 45.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Pacote com 15 sachês de alimento úmido Purina Friskies para gatos filhotes, sabor carne ao molho, 85 g cada sachê. Produto novo.",
        },
    )

    fill_row(
        ws,
        44,
        {
            "B": DATE,
            "C": "Livros e religião",
            "D": "Livro Café com Deus Pai, edição nova, porções diárias de amor, de Junior Rostirola.",
            "E": "Junior Rostirola",
            "F": "Café com Deus Pai",
            "G": "Novo",
            "H": "Não",
            "I": "Produto físico impresso; conferir capa, miolo, lombada e plástico externo.",
            "J": "Disponível",
            "K": "A1",
            "N": 28.90,
            "O": 39.99,
            "P": 72.00,
            "Q": "Shopee nacional: CAFÉ COM DEUS PAI 2026 | JUNIOR ROSTIROLA R$111,90; Livro Devocional Café com Deus Pai | Porções Diárias de Renovação R$32,99; CAFÉ COM DEUS PAI 2026 - DEVOCIONAL - JUNIOR ROSTIROLA R$39,90.",
            "S": 40.00,
            "T": 40.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Livro Café com Deus Pai, edição nova, porções diárias de amor, de Junior Rostirola. Produto novo e lacrado.",
        },
    )

    fill_row(
        ws,
        45,
        {
            "B": DATE,
            "C": "Livros e esoterismo",
            "D": "Livro O Caibalion, Os Três Iniciados, estudo da filosofia hermética do Egito antigo e da Grécia.",
            "E": "Camelot / Os Três Iniciados",
            "F": "O Caibalion",
            "G": "Novo",
            "H": "Não",
            "I": "Produto físico impresso; conferir capa, miolo, lombada e plástico externo.",
            "J": "Disponível",
            "K": "A1",
            "N": 18.90,
            "O": 22.90,
            "P": 35.00,
            "Q": "Shopee nacional: Livro O Caibalion - Camelot | Os Três Iniciados R$18,90; O Caibalion: As Sete Leis Universais... R$35,00; O Caibalion: Os Três Iniciados R$22,90.",
            "S": 20.00,
            "T": 20.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Livro O Caibalion, Os Três Iniciados, estudo da filosofia hermética do Egito antigo e da Grécia. Produto novo.",
        },
    )

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
