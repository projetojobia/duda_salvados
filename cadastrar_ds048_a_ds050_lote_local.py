from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
DATE = "10/08/2026"


def fill_row(ws, row: int, data: dict[str, object]) -> None:
    for col, value in data.items():
        ws[f"{col}{row}"] = value
    ws[f"AE{row}"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    fill_row(
        ws,
        49,
        {
            "B": DATE,
            "C": "Calçados",
            "D": "Tênis esportivo preto Dostin Hike, com sola em tom dourado e design casual/treino.",
            "E": "Dostin",
            "F": "Hike preto",
            "G": "Novo",
            "H": "Não",
            "I": "Produto de calçado; conferir numeração, acabamento, sola e costura.",
            "J": "Disponível",
            "K": "A1",
            "N": 119.90,
            "O": 149.90,
            "P": 199.90,
            "Q": "Shopee nacional: Tênis Dostin Hike preto masculino R$149,90; tênis casual preto com sola gum e detalhes dourados R$119,90; tênis esportivo preto estilo Hike R$199,90.",
            "S": 120.00,
            "T": 120.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Tênis esportivo preto Dostin Hike, com sola em tom dourado e design casual/treino. Produto novo.",
        },
    )

    fill_row(
        ws,
        50,
        {
            "B": DATE,
            "C": "Ferramentas e construção",
            "D": "Rolo de arame farpado em aço zincado, indicado para cercas e contenção, com resistência de 50 kgf.",
            "E": "Não identificada",
            "F": "Arame farpado aço zincado 50 kgf",
            "G": "Novo",
            "H": "Não",
            "I": "Produto de construção; conferir peso, estado do rolo e integridade do zincado.",
            "J": "Disponível",
            "K": "A1",
            "N": 79.90,
            "O": 99.90,
            "P": 159.90,
            "Q": "Shopee nacional: Arame Farpado Aço Zincado 50kgf rolo para cerca R$99,90; arame farpado galvanizado para cerca R$79,90; rolo de arame farpado zincado reforçado R$159,90.",
            "S": 100.00,
            "T": 100.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Rolo de arame farpado em aço zincado, indicado para cercas e contenção, com resistência de 50 kgf. Produto novo.",
        },
    )

    fill_row(
        ws,
        51,
        {
            "B": DATE,
            "C": "Ferramentas automotivas",
            "D": "Macaco hidráulico tipo jacaré Mtx, capacidade de 2 toneladas, para elevação de veículos.",
            "E": "Mtx",
            "F": "Macaco hidráulico 2T jacaré",
            "G": "Novo",
            "H": "Não",
            "I": "Ferramenta automotiva; conferir rodas, alavanca e funcionamento hidráulico antes da publicação.",
            "J": "Disponível",
            "K": "A1",
            "N": 189.90,
            "O": 249.90,
            "P": 349.90,
            "Q": "Shopee nacional: Macaco hidráulico tipo jacaré 2 toneladas R$249,90; macaco automotivo 2T baixo perfil R$189,90; macaco hidráulico 2 ton profissional R$349,90.",
            "S": 200.00,
            "T": 200.00,
            "W": "Não",
            "X": "Não publicado",
            "Y": "Macaco hidráulico tipo jacaré Mtx, capacidade de 2 toneladas, para elevação de veículos. Produto novo.",
        },
    )

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
