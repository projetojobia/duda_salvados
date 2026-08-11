from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
SHEET_NAME = "Produtos"
ROW = 34


def main() -> None:
    wb = load_workbook(WORKBOOK_PATH)
    ws = wb[SHEET_NAME]

    values = {
        "B": "10/08/2026",
        "C": "Casa e suporte de TV",
        "D": "Suporte de parede articulado para TV, compatível com telas de 32 a 75 polegadas, VESA 600x400, carga máxima de 35 kg.",
        "E": "Não identificada",
        "F": "Suporte articulado para TV 32-75",
        "G": "Novo",
        "H": "Não",
        "I": "Produto sem mecanismo eletrônico; conferir estrutura, articulações, parafusos e acabamento antes de publicar.",
        "J": "Disponível",
        "K": "A1",
        "N": 34.90,
        "O": 57.75,
        "P": 99.39,
        "Q": (
            "Shopee nacional: Suporte Articulado de Parede para TV 32 a 55 LCD/LED/OLED/QLED/Plasma R$34,90; "
            "Suporte Articulado para TV 32'' a 60'' Parede Giro Inclinação Aço Reforçado R$57,75; "
            "Suporte Tv 10 32 42 65 Articulado Parede Vesa 600x400 Branco R$99,39."
        ),
        "S": 55.00,
        "T": 75.00,
        "W": "Não",
        "X": "Não publicado",
        "Y": (
            "Suporte de parede articulado para TV, indicado para telas de 32 a 75 polegadas. Estrutura reforçada, "
            "com inclinação e ajuste lateral."
        ),
        "AE": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
    }

    for col, value in values.items():
        ws[f"{col}{ROW}"] = value

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    main()
