from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


WORKBOOK_PATH = Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
PRODUCTS_SHEET = "Produtos"

PRODUCT_HEADERS = [
    "Codigo",
    "Data cadastro",
    "Categoria",
    "Produto / descricao",
    "Marca",
    "Modelo",
    "Condicao",
    "Testado?",
    "Funcionamento",
    "Status interno",
    "Localizacao fisica",
    "Foto principal (URL)",
    "Pasta / fotos (URL)",
    "Preco baixo IA (R$)",
    "Preco medio IA (R$)",
    "Preco alto IA (R$)",
    "Fonte precos / URL",
    "Custo estimado ref. (R$)",
    "Preco sugerido IA (R$)",
    "Preco definido (R$)",
    "Vantagem ref. mercado (R$)",
    "Vantagem ref. mercado (%)",
    "Publicar no WhatsApp?",
    "Status catalogo",
    "Descricao para anuncio",
    "Data venda",
    "Preco venda real (R$)",
    "Lucro real ref. (R$)",
    "Cliente (opcional)",
    "Observacoes",
    "Ultima atualizacao",
]


@dataclass(frozen=True)
class ProductRow:
    code: str
    row_number: int
    values: list[Any]

    @property
    def is_free(self) -> bool:
        return bool(self.code) and all(_is_empty(value) for value in self.values[1:])

    def as_dict(self) -> dict[str, Any]:
        padded = _pad_row(self.values)
        return {PRODUCT_HEADERS[index]: padded[index] for index in range(len(PRODUCT_HEADERS))}


def _is_empty(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def _pad_row(row: list[Any], size: int = 31) -> list[Any]:
    padded = list(row[:size])
    while len(padded) < size:
        padded.append("")
    return padded


def _brl(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _parse_brl(value: Any) -> float | None:
    if _is_empty(value):
        return None
    text = str(value).replace("R$", "").replace(".", "").replace(",", ".")
    digits = []
    for ch in text:
        if ch.isdigit() or ch in ".-":
            digits.append(ch)
    cleaned = "".join(digits)
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_code(code: str) -> str:
    cleaned = "".join(ch for ch in code.upper() if ch.isalnum())
    if not cleaned.startswith("DS"):
        raise ValueError(f"Codigo invalido: {code!r}")
    return cleaned


class LocalWorkbookHelper:
    def __init__(self, path: Path = WORKBOOK_PATH) -> None:
        self.path = path

    def load(self):
        return load_workbook(self.path)

    def save(self, workbook) -> None:
        workbook.save(self.path)

    def list_product_rows(self, end_row: int = 301) -> list[ProductRow]:
        wb = self.load()
        ws = wb[PRODUCTS_SHEET]
        rows: list[ProductRow] = []
        for row_number in range(2, end_row + 1):
            values = [ws.cell(row=row_number, column=col).value for col in range(1, 32)]
            code = str(values[0] or "").strip()
            if code:
                rows.append(ProductRow(code=code, row_number=row_number, values=values))
        return rows

    def find_next_free_code(self, end_row: int = 301) -> ProductRow:
        for row in self.list_product_rows(end_row=end_row):
            if row.is_free:
                return row
        raise RuntimeError("Nenhum codigo DSxxx livre encontrado.")

    def read_product(self, code: str) -> ProductRow:
        normalized = normalize_code(code)
        for row in self.list_product_rows():
            if row.code.upper() == normalized:
                return row
        raise RuntimeError(f"Codigo nao encontrado: {normalized}")

    def update_product_fields(self, code: str, fields: dict[str, Any], update_timestamp: bool = True) -> ProductRow:
        wb = self.load()
        ws = wb[PRODUCTS_SHEET]
        before = self.read_product(code)
        row_number = before.row_number
        updates = dict(fields)
        if update_timestamp:
            updates["AE"] = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

        for key, value in updates.items():
            col = _resolve_col(key)
            if col == "A":
                raise RuntimeError("Coluna protegida contra escrita: A")
            ws[f"{col}{row_number}"] = value

        self.save(wb)
        return self.read_product(before.code)

    def define_price(self, code: str, price: float) -> ProductRow:
        row = self.read_product(code)
        updates = {"T": price}
        if str(row.values[9]).strip() == "Aguardando preço":
            updates["J"] = "Pronto para publicar"
        return self.update_product_fields(row.code, updates)

    def reserve(self, code: str, customer: str | None = None) -> ProductRow:
        row = self.read_product(code)
        if str(row.values[9]).strip() != "Disponível":
            raise RuntimeError(f"Produto {row.code} nao esta Disponivel.")
        updates: dict[str, Any] = {"J": "Reservado", "W": "Não"}
        if customer:
            updates["AC"] = customer
        return self.update_product_fields(row.code, updates)

    def sell(self, code: str, sale_price: float, customer: str | None = None) -> ProductRow:
        row = self.read_product(code)
        status = str(row.values[9]).strip()
        if status not in {"Disponível", "Reservado", "Publicado"}:
            raise RuntimeError(f"Produto {row.code} nao pode ser vendido com status atual: {status!r}")
        if not _is_empty(row.values[26]):
            raise RuntimeError(f"Produto {row.code} ja tem preco de venda real em AA.")
        updates: dict[str, Any] = {
            "J": "Vendido",
            "Z": datetime.now().strftime("%d/%m/%Y"),
            "AA": sale_price,
            "W": "Não",
        }
        if str(row.values[23]).strip() == "Publicado":
            updates["X"] = "Remover"
        if customer:
            updates["AC"] = customer
        return self.update_product_fields(row.code, updates)


def _resolve_col(key: str) -> str:
    key = key.strip()
    if re_full := key.upper():
        if re_full in {chr(ord("A") + i) for i in range(26)} | {"AA", "AB", "AC", "AD", "AE"}:
            return re_full
    mapping = {field: col for col, field in zip(
        ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD", "AE"],
        PRODUCT_HEADERS,
    )}
    if key in mapping:
        return mapping[key]
    raise RuntimeError(f"Campo/coluna invalido: {key!r}")


def main() -> int:
    parser = argparse.ArgumentParser(description="Helper local para a planilha Duda Salvados.")
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("next-free")

    read_parser = sub.add_parser("read")
    read_parser.add_argument("code")

    set_parser = sub.add_parser("set")
    set_parser.add_argument("code")
    set_parser.add_argument("field_or_col")
    set_parser.add_argument("value")

    update_parser = sub.add_parser("update-json")
    update_parser.add_argument("code")
    update_parser.add_argument("fields_json")

    define_parser = sub.add_parser("define-price")
    define_parser.add_argument("code")
    define_parser.add_argument("price", type=float)

    reserve_parser = sub.add_parser("reserve")
    reserve_parser.add_argument("code")
    reserve_parser.add_argument("customer", nargs="?")

    sell_parser = sub.add_parser("sell")
    sell_parser.add_argument("code")
    sell_parser.add_argument("sale_price", type=float)
    sell_parser.add_argument("customer", nargs="?")

    args = parser.parse_args()
    client = LocalWorkbookHelper()

    if args.command == "next-free":
        row = client.find_next_free_code()
        print(json.dumps({"code": row.code, "row_number": row.row_number, "values": row.as_dict()}, ensure_ascii=False, indent=2))
    elif args.command == "read":
        row = client.read_product(args.code)
        print(json.dumps({"code": row.code, "row_number": row.row_number, "values": row.as_dict()}, ensure_ascii=False, indent=2))
    elif args.command == "set":
        updated = client.update_product_fields(args.code, {args.field_or_col: args.value})
        print(json.dumps({"code": updated.code, "row_number": updated.row_number, "values": updated.as_dict()}, ensure_ascii=False, indent=2))
    elif args.command == "update-json":
        fields = json.loads(args.fields_json)
        updated = client.update_product_fields(args.code, fields)
        print(json.dumps({"code": updated.code, "row_number": updated.row_number, "values": updated.as_dict()}, ensure_ascii=False, indent=2))
    elif args.command == "define-price":
        updated = client.define_price(args.code, args.price)
        print(json.dumps({"code": updated.code, "row_number": updated.row_number, "values": updated.as_dict()}, ensure_ascii=False, indent=2))
    elif args.command == "reserve":
        updated = client.reserve(args.code, args.customer)
        print(json.dumps({"code": updated.code, "row_number": updated.row_number, "values": updated.as_dict()}, ensure_ascii=False, indent=2))
    elif args.command == "sell":
        updated = client.sell(args.code, args.sale_price, args.customer)
        print(json.dumps({"code": updated.code, "row_number": updated.row_number, "values": updated.as_dict()}, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
