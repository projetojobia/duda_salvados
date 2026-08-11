import csv
import json
import pathlib
import re
import unicodedata
from openpyxl import load_workbook


WORKBOOK = pathlib.Path(r"C:\Users\User\duda\Duda_Salvados_Hermes_GoogleSheets_v2.xlsx")
OUT = pathlib.Path(r"C:\Users\User\duda\fotos_organizadas_manifest.csv")
BASE_DIR = pathlib.Path(r"C:\Users\User\duda\Fotos_Organizadas")


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text[:70]


def main() -> None:
    wb = load_workbook(WORKBOOK, data_only=True)
    ws = wb["Produtos"]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=73, min_col=1, max_col=15, values_only=True):
        if not row or len(row) < 4 or not row[0]:
            continue
        code, _, category, desc, brand, model, condition, tested, working, status, location, *_rest = row[:15]
        if not desc:
            continue
        slug = slugify(desc.split(".")[0])
        rows.append(
            {
                "Codigo": code,
                "Categoria": category,
                "Marca": brand,
                "Modelo": model,
                "Condicao": condition,
                "Testado": tested,
                "Funcionamento": working,
                "Status": status,
                "Localizacao": location,
                "Descricao": desc,
                "Arquivo_padrao": f"{code}_{slug}_01.jpg",
            }
        )

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with OUT.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "Codigo",
                "Categoria",
                "Marca",
                "Modelo",
                "Condicao",
                "Testado",
                "Funcionamento",
                "Status",
                "Localizacao",
                "Descricao",
                "Arquivo_padrao",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)

    BASE_DIR.mkdir(parents=True, exist_ok=True)
    for row in rows:
        slug = row["Arquivo_padrao"].replace(".jpg", "")
        folder = BASE_DIR / slug
        folder.mkdir(parents=True, exist_ok=True)

    print(OUT)


if __name__ == "__main__":
    main()
